import logging
from io import BytesIO
from typing import Any

from openpyxl import load_workbook

logger = logging.getLogger(__name__)

REQUIRED_FIELDS = ["char_no", "requirement", "reference_location"]


def _matches_field(cell_text: str, field: str) -> bool:
    """
    Field-specific header matching — mirrors the compound logic in load_form3.

    Uses precise compound rules instead of any-keyword matching to avoid false
    positives (e.g. "no" matching "zone no", "detail no", etc.).
    """
    t = cell_text.lower().strip()
    if field == "char_no":
        return "char" in t and ("no" in t or "number" in t)
    if field == "requirement":
        return "requirement" in t or (t == "req")
    if field == "reference_location":
        return "reference" in t or "location" in t
    if field == "characteristic_designator":
        return "characteristic" in t or "designator" in t
    return False


def _find_header_row(
    rows: list[tuple[Any, ...]],
) -> tuple[int, list[str | None], dict[str, int]] | None:
    """
    Scan all rows to locate the AS9102 Form 3 header row.

    Mirrors the scanning logic in delta_preservation/io/xlsx.py:load_form3 so the
    web preview uses the same detection as the pipeline.  Returns (row_index,
    headers, detected_mapping) for the first row with ≥ 2 recognised field headers,
    or None if no such row is found.
    """
    all_fields = ["char_no", "requirement", "reference_location", "characteristic_designator"]

    for row_idx, row in enumerate(rows):
        col_map: dict[str, int] = {}
        for col_idx, cell in enumerate(row):
            if cell is None:
                continue
            cell_text = str(cell)
            for field in all_fields:
                if field not in col_map and _matches_field(cell_text, field):
                    col_map[field] = col_idx
        if len(col_map) >= 2:
            headers = [str(h) if h is not None else None for h in row]
            return row_idx, headers, col_map

    return None


def detect_column_mapping(headers: list[str | None]) -> dict[str, int]:
    """
    Auto-detect column indices for required fields from a known header row.

    Uses field-specific compound matching to avoid false positives.
    Returns {field_name: col_index} for confident matches only.
    Unmatched fields are omitted — caller highlights them as amber.
    """
    detected: dict[str, int] = {}
    for field in REQUIRED_FIELDS:
        for col_idx, header in enumerate(headers):
            if header is None:
                continue
            if _matches_field(str(header), field):
                detected[field] = col_idx
                break
    return detected


def parse_excel_preview(
    file_bytes: bytes,
) -> tuple[list[str | None], list[tuple[Any, ...]], dict[str, int]]:
    """
    Parse Excel bytes for column mapping preview and auto-detection.

    Scans all rows to find the AS9102 Form 3 header row (same logic as the
    pipeline's load_form3), so files with title/metadata rows above the headers
    are handled correctly.

    Returns: (headers, preview_rows[:5], detected_mapping)
    Raises ValueError with a descriptive message on fatal errors (empty file,
    unreadable bytes, sheet with no rows, no header row found).
    Non-fatal issues (columns not auto-detected) result in those fields being
    omitted from detected_mapping — not raised as errors.
    """
    if not file_bytes:
        raise ValueError("File is empty — please upload a non-empty Excel file.")

    try:
        wb = load_workbook(BytesIO(file_bytes), read_only=True, data_only=True)
    except Exception as exc:
        raise ValueError(f"Cannot read file: {exc}") from exc

    ws = wb.active
    rows = list(ws.iter_rows(values_only=True))

    if not rows:
        raise ValueError("Sheet has no rows — please upload a file with data.")

    result = _find_header_row(rows)
    if result is None:
        # Fall back to row 0 so the user can still manually map columns.
        logger.warning("parse_excel_preview: no AS9102 header row found — falling back to row 0")
        headers = [str(h) if h is not None else None for h in rows[0]]
        preview_rows = list(rows[1:6])
        detected: dict[str, int] = {}
    else:
        header_row_idx, headers, detected_all = result
        preview_rows = list(rows[header_row_idx + 1 : header_row_idx + 6])
        # Filter to REQUIRED_FIELDS only for the UI dropdowns.
        detected = {f: col for f, col in detected_all.items() if f in REQUIRED_FIELDS}

    logger.info(
        "parse_excel_preview: cols=%d, detected=%s",
        len(headers),
        list(detected.keys()),
    )
    return headers, preview_rows, detected
