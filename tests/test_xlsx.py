import pytest
from pathlib import Path
import json
from openpyxl import Workbook
from delta_preservation.io.xlsx import load_form3, Characteristic


@pytest.fixture
def temp_intermediate_dir(tmp_path):
    """Create a temporary intermediate directory for debug output."""
    intermediate = tmp_path / "intermediate"
    intermediate.mkdir()
    return intermediate


def create_test_workbook(sheet_name: str = "Form3") -> Workbook:
    """Helper to create a basic test workbook with Form3 sheet."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    return wb


def add_headers_row6(ws, use_keywords=True):
    """Add header row at row 6."""
    if use_keywords:
        ws.cell(row=6, column=1, value="Char No")
        ws.cell(row=6, column=2, value="Reference Location")
        ws.cell(row=6, column=3, value="Characteristic Designator")
        ws.cell(row=6, column=4, value="Requirement")
    else:
        # Non-matching headers
        ws.cell(row=6, column=1, value="Column A")
        ws.cell(row=6, column=2, value="Column B")
        ws.cell(row=6, column=3, value="Column C")
        ws.cell(row=6, column=4, value="Column D")


def add_characteristic_row(ws, row_num: int, char_no, ref_loc, designator, requirement):
    """Add a characteristic row to the worksheet."""
    ws.cell(row=row_num, column=1, value=char_no)
    ws.cell(row=row_num, column=2, value=ref_loc)
    ws.cell(row=row_num, column=3, value=designator)
    ws.cell(row=row_num, column=4, value=requirement)


# ============================================================================
# 1) Basic "happy path" parse works
# ============================================================================

def test_returns_list(tmp_path, temp_intermediate_dir):
    """Test that load_form3 returns a list."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert isinstance(result, list)


def test_list_non_empty_for_known_good_fixture(tmp_path, temp_intermediate_dir):
    """Test that a known-good fixture returns a non-empty list."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, 2, "SHEET 2", "DIM B", "20.0 ± 0.2")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) > 0


def test_all_elements_are_characteristic_objects(tmp_path, temp_intermediate_dir):
    """Test that all elements are Characteristic objects with expected attributes."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, 2, "SHEET 2", "DIM B", "20.0 ± 0.2")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    
    for char in result:
        assert isinstance(char, Characteristic)
        assert hasattr(char, 'char_no')
        assert hasattr(char, 'reference_location')
        assert hasattr(char, 'characteristic_designator')
        assert hasattr(char, 'requirement')
        assert isinstance(char.char_no, int)
        assert isinstance(char.reference_location, str)
        assert isinstance(char.characteristic_designator, str)
        assert isinstance(char.requirement, str)


# ============================================================================
# 2) It reads the right sheet, and errors clearly when missing
# ============================================================================

def test_parses_sheet_named_f3(tmp_path, temp_intermediate_dir):
    """Test that a sheet named 'F3' is parsed correctly."""
    wb = create_test_workbook(sheet_name="F3")
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 1


def test_parses_sheet_with_form3_substring(tmp_path, temp_intermediate_dir):
    """Test that a sheet with 'Form3' substring (e.g., 'MyForm3') is parsed."""
    wb = create_test_workbook(sheet_name="MyForm3Data")
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 1


def test_parses_sheet_with_f3_substring(tmp_path, temp_intermediate_dir):
    """Test that a sheet with 'F3' substring is parsed."""
    wb = create_test_workbook(sheet_name="DataF3Sheet")
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 1


def test_raises_valueerror_when_form3_sheet_missing(tmp_path, temp_intermediate_dir):
    """Test that ValueError is raised when neither Form3 nor F3 sheet exists."""
    wb = Workbook()
    ws = wb.active
    ws.title = "SomeOtherSheet"
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    with pytest.raises(ValueError) as exc_info:
        load_form3(xlsx_path, temp_intermediate_dir)
    
    assert "Form3" in str(exc_info.value) or "F3" in str(exc_info.value)


# ============================================================================
# 3) It starts at the correct row and stops correctly
# ============================================================================

def test_starts_at_row_7(tmp_path, temp_intermediate_dir):
    """Test that parsing starts at row 7 (ignores earlier rows)."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    # Add data in rows 1-6 (should be ignored)
    add_characteristic_row(ws, 1, 99, "IGNORED", "IGNORED", "IGNORED")
    add_characteristic_row(ws, 5, 98, "IGNORED", "IGNORED", "IGNORED")
    # Add valid data starting at row 7
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, 2, "SHEET 2", "DIM B", "20.0 ± 0.2")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 2
    assert result[0].char_no == 1
    assert result[1].char_no == 2


def test_stops_when_char_no_is_blank(tmp_path, temp_intermediate_dir):
    """Test that parsing stops when Char No cell is empty."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, 2, "SHEET 2", "DIM B", "20.0 ± 0.2")
    # Row 9 has blank Char No
    add_characteristic_row(ws, 9, None, "SHEET 3", "DIM C", "30.0 ± 0.3")
    # Row 10 has data but should be ignored
    add_characteristic_row(ws, 10, 4, "SHEET 4", "DIM D", "40.0 ± 0.4")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 2


def test_empty_result_when_row_7_char_no_is_blank(tmp_path, temp_intermediate_dir):
    """Test that result is empty when row 7 Char No is blank."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, None, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, 1, "SHEET 2", "DIM B", "20.0 ± 0.2")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 0


def test_stops_at_blank_row_ignores_data_after(tmp_path, temp_intermediate_dir):
    """Test that a blank row in the middle stops parsing (limitation)."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, 2, "SHEET 2", "DIM B", "20.0 ± 0.2")
    # Row 9 is blank
    add_characteristic_row(ws, 9, None, None, None, None)
    # Row 10 has data but will be ignored due to current implementation
    add_characteristic_row(ws, 10, 3, "SHEET 3", "DIM C", "30.0 ± 0.3")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    # Current behavior: stops at blank row, ignores row 10
    assert len(result) == 2
    assert result[0].char_no == 1
    assert result[1].char_no == 2


# ============================================================================
# 4) Char No parsing and skipping behavior
# ============================================================================

def test_char_no_as_integer(tmp_path, temp_intermediate_dir):
    """Test that Char No as integer is parsed correctly."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 7, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 1
    assert result[0].char_no == 7


def test_char_no_as_string_integer(tmp_path, temp_intermediate_dir):
    """Test that Char No as string '7' is parsed as integer 7."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, "7", "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 1
    assert result[0].char_no == 7


def test_char_no_as_float(tmp_path, temp_intermediate_dir):
    """Test that Char No as float 7.0 is parsed as integer 7."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 7.0, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 1
    assert result[0].char_no == 7


def test_char_no_invalid_string_skipped(tmp_path, temp_intermediate_dir):
    """Test that rows with invalid Char No strings are skipped."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, "7A", "SHEET 2", "DIM B", "20.0 ± 0.2")  # Invalid
    add_characteristic_row(ws, 9, "A7", "SHEET 3", "DIM C", "30.0 ± 0.3")  # Invalid
    add_characteristic_row(ws, 10, 2, "SHEET 4", "DIM D", "40.0 ± 0.4")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 2
    assert result[0].char_no == 1
    assert result[1].char_no == 2


def test_char_no_with_spaces_parsed(tmp_path, temp_intermediate_dir):
    """Test that Char No with spaces '  7  ' is parsed correctly."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, "  7  ", "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 1
    assert result[0].char_no == 7


def test_char_nos_are_unique(tmp_path, temp_intermediate_dir):
    """Test that Char Nos are unique in output (detects duplicates)."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, 2, "SHEET 2", "DIM B", "20.0 ± 0.2")
    add_characteristic_row(ws, 9, 3, "SHEET 3", "DIM C", "30.0 ± 0.3")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    char_nos = [char.char_no for char in result]
    assert len(char_nos) == len(set(char_nos)), "Duplicate Char Nos detected"


def test_duplicate_char_nos_detected(tmp_path, temp_intermediate_dir):
    """Test that duplicate Char Nos are detected (negative test)."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, 2, "SHEET 2", "DIM B", "20.0 ± 0.2")
    add_characteristic_row(ws, 9, 1, "SHEET 3", "DIM C", "30.0 ± 0.3")  # Duplicate
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    char_nos = [char.char_no for char in result]
    # This test documents that duplicates ARE allowed by current implementation
    assert len(char_nos) != len(set(char_nos)), "Expected duplicate Char Nos"


# ============================================================================
# 5) Column detection logic behaves as expected (and fallback works)
# ============================================================================

def test_header_detection_with_matching_keywords(tmp_path, temp_intermediate_dir):
    """Test that header detection succeeds when headers match keywords."""
    wb = create_test_workbook()
    ws = wb.active
    # Headers with all keywords present
    ws.cell(row=6, column=1, value="Char No Number")
    ws.cell(row=6, column=2, value="Reference Location Ref")
    ws.cell(row=6, column=3, value="Characteristic Designator")
    ws.cell(row=6, column=4, value="Requirement Req")
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 1
    assert result[0].char_no == 1
    assert result[0].reference_location == "SHEET 1"
    assert result[0].characteristic_designator == "DIM A"
    assert result[0].requirement == "10.0 ± 0.1"


def test_fallback_mapping_with_missing_headers(tmp_path, temp_intermediate_dir):
    """Test that fallback column mapping works when headers don't match."""
    wb = create_test_workbook()
    ws = wb.active
    # Non-matching headers
    ws.cell(row=6, column=1, value="Column A")
    ws.cell(row=6, column=2, value="Column B")
    ws.cell(row=6, column=3, value="Column C")
    ws.cell(row=6, column=4, value="Column D")
    # Data in fallback positions (A, B, C, D)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, 2, "SHEET 2", "DIM B", "20.0 ± 0.2")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 2
    assert result[0].char_no == 1
    assert result[0].reference_location == "SHEET 1"
    assert result[0].characteristic_designator == "DIM A"
    assert result[0].requirement == "10.0 ± 0.1"


def test_fallback_mapping_with_no_headers(tmp_path, temp_intermediate_dir):
    """Test that fallback works when row 6 is completely empty."""
    wb = create_test_workbook()
    ws = wb.active
    # Row 6 is empty (no headers)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert len(result) == 1
    assert result[0].char_no == 1
    assert result[0].reference_location == "SHEET 1"


# ============================================================================
# 6) String cleanup rules
# ============================================================================

def test_reference_location_stripped(tmp_path, temp_intermediate_dir):
    """Test that reference_location with spaces is stripped."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "  SHEET 1  ", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert result[0].reference_location == "SHEET 1"


def test_none_values_become_empty_strings(tmp_path, temp_intermediate_dir):
    """Test that None values are converted to empty strings."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, None, None, None)
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    assert result[0].reference_location == ""
    assert result[0].characteristic_designator == ""
    assert result[0].requirement == ""


def test_at_least_one_non_empty_requirement(tmp_path, temp_intermediate_dir):
    """Test that at least one row has a non-empty requirement in typical data."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, 2, "SHEET 2", "DIM B", "20.0 ± 0.2")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    non_empty_requirements = [char for char in result if char.requirement != ""]
    assert len(non_empty_requirements) >= 1


# ============================================================================
# 7) Debug JSON output behavior
# ============================================================================

def test_debug_json_file_created(tmp_path, temp_intermediate_dir):
    """Test that debug JSON file is created."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    load_form3(xlsx_path, temp_intermediate_dir)
    
    debug_json_path = temp_intermediate_dir / "form3_chars.json"
    assert debug_json_path.exists()


def test_debug_json_is_valid(tmp_path, temp_intermediate_dir):
    """Test that debug JSON is valid and can be loaded."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    load_form3(xlsx_path, temp_intermediate_dir)
    
    debug_json_path = temp_intermediate_dir / "form3_chars.json"
    with open(debug_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    assert isinstance(data, list)


def test_debug_json_entry_count_matches_result(tmp_path, temp_intermediate_dir):
    """Test that number of JSON entries equals number of returned characteristics."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, 2, "SHEET 2", "DIM B", "20.0 ± 0.2")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    
    debug_json_path = temp_intermediate_dir / "form3_chars.json"
    with open(debug_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    assert len(data) == len(result)


def test_debug_json_contains_expected_keys(tmp_path, temp_intermediate_dir):
    """Test that JSON objects contain expected keys."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    load_form3(xlsx_path, temp_intermediate_dir)
    
    debug_json_path = temp_intermediate_dir / "form3_chars.json"
    with open(debug_json_path, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    for entry in data:
        assert "char_no" in entry
        assert "reference_location" in entry
        assert "characteristic_designator" in entry
        assert "requirement" in entry


def test_intermediate_dir_must_exist(tmp_path):
    """Test that function fails if intermediate_dir doesn't exist."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    nonexistent_dir = tmp_path / "nonexistent"
    
    with pytest.raises(FileNotFoundError):
        load_form3(xlsx_path, nonexistent_dir)


# ============================================================================
# 8) Robustness / failure modes
# ============================================================================

def test_nonexistent_xlsx_path_raises(temp_intermediate_dir):
    """Test that nonexistent xlsx path raises an exception."""
    nonexistent_path = Path("/nonexistent/path/to/file.xlsx")
    
    with pytest.raises(Exception):
        load_form3(nonexistent_path, temp_intermediate_dir)


def test_invalid_char_no_rows_skipped_output_length_smaller(tmp_path, temp_intermediate_dir):
    """Test that rows with invalid Char No are skipped, reducing output length."""
    wb = create_test_workbook()
    ws = wb.active
    add_headers_row6(ws)
    add_characteristic_row(ws, 7, 1, "SHEET 1", "DIM A", "10.0 ± 0.1")
    add_characteristic_row(ws, 8, "INVALID", "SHEET 2", "DIM B", "20.0 ± 0.2")
    add_characteristic_row(ws, 9, "ALSO_INVALID", "SHEET 3", "DIM C", "30.0 ± 0.3")
    add_characteristic_row(ws, 10, 2, "SHEET 4", "DIM D", "40.0 ± 0.4")
    
    xlsx_path = tmp_path / "test.xlsx"
    wb.save(xlsx_path)
    
    result = load_form3(xlsx_path, temp_intermediate_dir)
    # Only 2 valid rows (rows 7 and 10), rows 8 and 9 skipped
    assert len(result) == 2
    assert result[0].char_no == 1
    assert result[1].char_no == 2
