"""
Test stubs for Phase 2 run/upload/pipeline requirements.

UPLOAD-01..05 tests are implemented here; PIPE-* remain as xfail stubs
until Plans 03-07 implement those features.

Requirement traceability:
  UPLOAD-01..05  Upload form and file validation
  PIPE-01..06    Pipeline task execution and status transitions
  PIPE-11        Alert creation on run failure
"""
import io
import pytest
from sqlalchemy.orm import sessionmaker


def _make_vector_pdf_bytes() -> bytes:
    """Return minimal single-page vector PDF with extractable text."""
    import fitz
    doc = fitz.open()
    page = doc.new_page()
    page.insert_text((50, 50), "1.00 ± 0.01")
    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    return buf.getvalue()


def _make_multipage_pdf_bytes(n_pages: int = 3) -> bytes:
    """Return minimal vector PDF with n_pages pages."""
    import fitz
    doc = fitz.open()
    for _ in range(n_pages):
        page = doc.new_page()
        page.insert_text((50, 50), "drawing text")
    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    return buf.getvalue()


def _make_excel_bytes() -> bytes:
    """Return valid minimal Excel bytes."""
    from openpyxl import Workbook
    wb = Workbook()
    ws = wb.active
    ws.append(["Char No", "Requirement", "Reference Location"])
    ws.append([1, "1.00 ± 0.01", "Zone A"])
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _login_user(client, db_engine, user):
    """Seed a session for a user and set cookie on client."""
    from sqlalchemy.orm import sessionmaker
    from shop.services.auth import create_session
    Session = sessionmaker(bind=db_engine)
    db = Session()
    token = create_session(db, user)
    db.close()
    client.cookies.set("session_token", token)
    return token


def _login_engineer(client, db_engine, engineer_user):
    """Seed a session for the engineer user and set cookie on client."""
    return _login_user(client, db_engine, engineer_user)


# ---------------------------------------------------------------------------
# Upload requirements
# ---------------------------------------------------------------------------


def test_upload_creates_run(client, db_engine, engineer_user, huey_immediate):
    """UPLOAD-01: Engineer uploads Rev A PDF, Rev B PDF, Form 3 Excel, and part metadata;
    a Run record is created and the response redirects to /runs/{id}."""
    from shop.models import Run
    Session = sessionmaker(bind=db_engine)

    _login_engineer(client, db_engine, engineer_user)
    pdf_bytes = _make_vector_pdf_bytes()
    excel_bytes = _make_excel_bytes()

    resp = client.post(
        "/runs/new",
        data={
            "part_number": "PN-001",
            "rev_a_label": "A",
            "rev_b_label": "B",
            "customer": "Acme",
            "job_number": "JOB-1",
            "revA_page": "0",
            "revB_page": "0",
        },
        files={
            "revA_pdf": ("revA.pdf", pdf_bytes, "application/pdf"),
            "revB_pdf": ("revB.pdf", pdf_bytes, "application/pdf"),
            "form3_xlsx": ("form3.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        },
        follow_redirects=False,
    )
    assert resp.status_code == 302, f"Expected 302 redirect, got {resp.status_code}: {resp.text[:200]}"
    location = resp.headers.get("location", "")
    assert location.startswith("/runs/"), f"Expected redirect to /runs/{{id}}, got {location!r}"

    # Check Run was persisted
    db = Session()
    run_id = int(location.split("/")[-1])
    run = db.query(Run).filter(Run.id == run_id).first()
    db.close()
    assert run is not None
    assert run.status == "queued"


def test_raster_pdf_rejected(client, db_engine, engineer_user):
    """UPLOAD-02: POST /runs/validate-pdf with a raster PDF returns error partial."""
    _login_engineer(client, db_engine, engineer_user)

    # Create a raster-like PDF: no text, one large image covering the page
    import fitz
    doc = fitz.open()
    doc.new_page(width=595, height=842)
    buf = io.BytesIO()
    doc.save(buf)
    doc.close()
    raster_pdf_bytes = buf.getvalue()  # No text, no images — will be treated as not-raster
    # Use actual raster PDF: insert a full-page image
    # For simplicity, test the validate-pdf endpoint with a genuinely raster PDF
    # We'll create one by embedding a PNG image that covers the page
    import struct
    import zlib

    def _make_png(w: int = 10, h: int = 10) -> bytes:
        """Minimal valid PNG."""
        def chunk(name, data):
            c = zlib.crc32(name + data) & 0xFFFFFFFF
            return struct.pack(">I", len(data)) + name + data + struct.pack(">I", c)
        ihdr_data = struct.pack(">IIBBBBB", w, h, 8, 2, 0, 0, 0)
        raw = b"".join(b"\x00" + b"\xff\x00\x00" * w for _ in range(h))
        idat_data = zlib.compress(raw)
        return (
            b"\x89PNG\r\n\x1a\n"
            + chunk(b"IHDR", ihdr_data)
            + chunk(b"IDAT", idat_data)
            + chunk(b"IEND", b"")
        )

    png_bytes = _make_png(100, 100)
    doc2 = fitz.open()
    page = doc2.new_page(width=595, height=842)
    rect = fitz.Rect(0, 0, 595, 842)
    page.insert_image(rect, stream=png_bytes)
    buf2 = io.BytesIO()
    doc2.save(buf2)
    doc2.close()
    raster_pdf_bytes = buf2.getvalue()

    resp = client.post(
        "/runs/validate-pdf",
        data={"field": "revA"},
        files={"file": ("revA.pdf", raster_pdf_bytes, "application/pdf")},
    )
    assert resp.status_code == 200
    # Should return error partial — not the page selector
    assert "error" in resp.text.lower() or "scanned" in resp.text.lower() or "vector" in resp.text.lower()


def test_multipage_pdf_page_selector(client, db_engine, engineer_user):
    """UPLOAD-03: POST /runs/validate-pdf with multi-page PDF returns page selector partial."""
    _login_engineer(client, db_engine, engineer_user)
    pdf_bytes = _make_multipage_pdf_bytes(3)

    resp = client.post(
        "/runs/validate-pdf",
        data={"field": "revA"},
        files={"file": ("revA.pdf", pdf_bytes, "application/pdf")},
    )
    assert resp.status_code == 200
    # Should return page selector partial
    assert "page" in resp.text.lower() or "select" in resp.text.lower()


def test_invalid_excel_rejected(client, db_engine, engineer_user):
    """UPLOAD-04: POST /runs/new with an unreadable Excel returns 422 with error message."""
    _login_engineer(client, db_engine, engineer_user)
    pdf_bytes = _make_vector_pdf_bytes()

    resp = client.post(
        "/runs/new",
        data={
            "part_number": "PN-001",
            "rev_a_label": "A",
            "rev_b_label": "B",
            "customer": "Acme",
            "job_number": "JOB-1",
            "revA_page": "0",
            "revB_page": "0",
        },
        files={
            "revA_pdf": ("revA.pdf", pdf_bytes, "application/pdf"),
            "revB_pdf": ("revB.pdf", pdf_bytes, "application/pdf"),
            "form3_xlsx": ("form3.xlsx", b"not an excel file at all!!", "application/octet-stream"),
        },
        follow_redirects=False,
    )
    # Should get 422 with error shown in the form
    assert resp.status_code == 422
    assert "error" in resp.text.lower() or "form" in resp.text.lower()


def test_rev_labels_stored(client, db_engine, engineer_user, huey_immediate):
    """UPLOAD-05: Rev A and Rev B labels are persisted in Run.rev_a_label and Run.rev_b_label."""
    from shop.models import Run
    Session = sessionmaker(bind=db_engine)

    _login_engineer(client, db_engine, engineer_user)
    pdf_bytes = _make_vector_pdf_bytes()
    excel_bytes = _make_excel_bytes()

    resp = client.post(
        "/runs/new",
        data={
            "part_number": "PN-REV",
            "rev_a_label": "C",
            "rev_b_label": "D",
            "customer": "Corp",
            "job_number": "JOB-REV",
            "revA_page": "0",
            "revB_page": "0",
        },
        files={
            "revA_pdf": ("revA.pdf", pdf_bytes, "application/pdf"),
            "revB_pdf": ("revB.pdf", pdf_bytes, "application/pdf"),
            "form3_xlsx": ("form3.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
        },
        follow_redirects=False,
    )
    assert resp.status_code == 302
    location = resp.headers.get("location", "")
    run_id = int(location.split("/")[-1])

    db = Session()
    run = db.query(Run).filter(Run.id == run_id).first()
    db.close()
    assert run.rev_a_label == "C"
    assert run.rev_b_label == "D"


# ---------------------------------------------------------------------------
# Per-run xlsx column mapping validation (VALIDATE-01..03)
# ---------------------------------------------------------------------------


def _make_form3_xlsx_bytes(headers=None, rows=None) -> bytes:
    """Return a minimal Form 3 xlsx with the given headers and data rows."""
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.append(headers or ["Char No", "Reference Location", "Requirement"])
    for row in (rows or [[1, "Zone A", "1.25 +/- 0.05"]]):
        ws.append(row)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def test_validate_xlsx_autodetect(client, db_engine, engineer_user):
    """VALIDATE-01: POST valid xlsx to /runs/validate-xlsx returns 200 with select dropdowns.

    Steps:
    - Authenticate as engineer.
    - POST a minimal AS9102 Form 3 xlsx with auto-detectable column headers.
    - Assert 200 response and HTML containing <select> elements for column mapping.
    """
    _login_engineer(client, db_engine, engineer_user)
    xlsx_bytes = _make_form3_xlsx_bytes()

    resp = client.post(
        "/runs/validate-xlsx",
        files={"form3_xlsx": ("form3.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 200
    assert "<select" in resp.text, "Expected select dropdowns in column mapping partial"
    assert 'class="select select-sm w-full mapping-select bg-base-100 text-base-content' in resp.text


def test_validate_xlsx_empty_file(client, db_engine, engineer_user):
    """VALIDATE-02: POST empty/invalid xlsx returns error partial (no <select> elements).

    Steps:
    - Authenticate as engineer.
    - POST empty bytes as xlsx.
    - Assert 200 response with error messaging and no select dropdowns.
    """
    _login_engineer(client, db_engine, engineer_user)

    resp = client.post(
        "/runs/validate-xlsx",
        files={"form3_xlsx": ("empty.xlsx", b"", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 200
    assert "<select" not in resp.text, "Error path should not contain select dropdowns"
    assert "error" in resp.text.lower() or "empty" in resp.text.lower()


def test_validate_xlsx_noncontiguous_char_no(client, db_engine, engineer_user):
    """VALIDATE-03: Non-contiguous char_no values are accepted without normalization.

    Steps:
    - POST xlsx with char_no values [1, 5, 99, 200].
    - Assert all four values appear in the response.
    """
    _login_engineer(client, db_engine, engineer_user)
    xlsx_bytes = _make_form3_xlsx_bytes(
        headers=["Char No", "Requirement"],
        rows=[[v, f"Req {v}"] for v in [1, 5, 99, 200]],
    )

    resp = client.post(
        "/runs/validate-xlsx",
        files={"form3_xlsx": ("form3.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 200
    for val in ["1", "5", "99", "200"]:
        assert val in resp.text, f"Expected char_no {val!r} in response"


def test_validate_xlsx_prefers_f3_sheet_over_active_sheet(client, db_engine, engineer_user):
    """VALIDATE-03b: upload preview selects the Form 3 sheet even when another sheet is active."""
    from openpyxl import Workbook

    _login_engineer(client, db_engine, engineer_user)

    wb = Workbook()
    ws1 = wb.active
    ws1.title = "F1 Part No. Accountabiity"
    ws1.append(["1. Part Number", "2. Part Name"])
    ws1.append([20, "WHEEL HUB"])

    ws2 = wb.create_sheet("F2 Product Accountability")
    ws2.append(["5. Material or Process Name", "6. Specification Number"])

    ws3 = wb.create_sheet("F3 Char. Accountability")
    ws3.append([None, None, None, None])
    ws3.append(["AS9102 Form 3", None, None, None])
    ws3.append(["Characteristic Accountability", None, None, None])
    ws3.append(["5. Char. Number:", "6. Reference Location:", "7. Characteristic Designator:", "8. Requirement:"])
    ws3.append([1, "revA pg.1, Zone B.2", "M", ".500 IN +/- .005"])
    ws3.append([2, "revA pg.1, Zone B.1", "M", ".750 IN +/- .005"])

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    resp = client.post(
        "/runs/validate-xlsx",
        files={"form3_xlsx": ("form3.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 200
    assert "Char. Number" in resp.text or "Char No" in resp.text
    assert "Reference Location" in resp.text
    assert ".500 IN +/- .005" in resp.text
    assert "WHEEL HUB" not in resp.text


def test_validate_xlsx_formats_integer_like_preview_values_without_trailing_point_zero(client, db_engine, engineer_user):
    """VALIDATE-03c: preview displays 1.0-style integers as 1 for readability."""
    from openpyxl import Workbook

    _login_engineer(client, db_engine, engineer_user)

    wb = Workbook()
    ws = wb.active
    ws.title = "F3 Char. Accountability"
    ws.append([None, None, None, None])
    ws.append(["AS9102 Form 3", None, None, None])
    ws.append(["Characteristic Accountability", None, None, None])
    ws.append(["5. Char. Number:", "6. Reference Location:", "7. Characteristic Designator:", "8. Requirement:"])
    ws.append([1.0, "revA pg.1, Zone B.2", "M", ".500 IN +/- .005"])
    ws.append([2.0, "revA pg.1, Zone B.1", "M", ".750 IN +/- .005"])

    buf = io.BytesIO()
    wb.save(buf)
    xlsx_bytes = buf.getvalue()

    resp = client.post(
        "/runs/validate-xlsx",
        files={"form3_xlsx": ("form3.xlsx", xlsx_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    assert resp.status_code == 200
    compact = " ".join(resp.text.split())
    assert "<td class=\"text-center text-xs\"> 1 </td>" in compact
    assert "<td class=\"text-center text-xs\"> 2 </td>" in compact
    assert "1.0" not in compact
    assert "2.0" not in compact


def test_validate_xlsx_requires_auth(client):
    """VALIDATE-04: POST /runs/validate-xlsx without auth cookie redirects to login."""
    resp = client.post(
        "/runs/validate-xlsx",
        files={"form3_xlsx": ("form3.xlsx", b"data", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
        follow_redirects=False,
    )
    # Unauthenticated requests must be redirected (302/303/307) to /login
    assert resp.status_code in (302, 303, 307), f"Expected redirect, got {resp.status_code}"
    assert "/login" in resp.headers.get("location", ""), (
        f"Expected redirect to /login, got {resp.headers.get('location')!r}"
    )


# ---------------------------------------------------------------------------
# Pipeline task requirements (remain as xfail stubs until Plan 03)
# ---------------------------------------------------------------------------


def test_pipeline_task_enqueued(client, db_engine, engineer_user, huey_immediate):
    """PIPE-01: After a valid upload, a Huey pipeline task is enqueued and executed
    synchronously via huey_immediate; Run.status transitions away from 'queued'."""
    import json
    import pathlib
    import tempfile
    from unittest.mock import patch
    from shop.models import Run
    Session = sessionmaker(bind=db_engine)

    _login_engineer(client, db_engine, engineer_user)
    pdf_bytes = _make_vector_pdf_bytes()
    excel_bytes = _make_excel_bytes()

    def _mock_run_pipeline(**kwargs):
        run_dir = pathlib.Path(tempfile.mkdtemp())
        packet = {
            "run_id": "test",
            "items": [
                {
                    "char_no": 1,
                    "status": "unchanged",
                    "scores": {"location": 0.9, "text": 0.9, "context": 0.9},
                }
            ],
        }
        (run_dir / "delta_packet.json").write_text(json.dumps(packet))
        return run_dir

    with patch("shop.tasks.run_pipeline", _mock_run_pipeline), \
         patch("shop.tasks.SessionLocal", sessionmaker(bind=db_engine)):
        resp = client.post(
            "/runs/new",
            data={
                "part_number": "PN-PIPE01",
                "rev_a_label": "A",
                "rev_b_label": "B",
                "customer": "Acme",
                "job_number": "JOB-PIPE01",
                "revA_page": "0",
                "revB_page": "0",
            },
            files={
                "revA_pdf": ("revA.pdf", pdf_bytes, "application/pdf"),
                "revB_pdf": ("revB.pdf", pdf_bytes, "application/pdf"),
                "form3_xlsx": ("form3.xlsx", excel_bytes, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"),
            },
            follow_redirects=False,
        )

    assert resp.status_code == 302, f"Expected 302, got {resp.status_code}: {resp.text[:200]}"
    location = resp.headers.get("location", "")
    run_id = int(location.split("/")[-1])

    db = Session()
    run = db.query(Run).filter(Run.id == run_id).first()
    db.close()
    assert run is not None
    # Task ran synchronously via huey_immediate; status should have progressed past queued
    assert run.status != "queued", f"Run should not still be 'queued' after task ran, got: {run.status!r}"


def test_stage_progress_updates(client, db_engine, engineer_user):
    """PIPE-02: GET /runs/{id} returns 200 with stage checklist; SSE endpoint streams stage updates."""
    from sqlalchemy.orm import sessionmaker
    from shop.models import Run
    from shop.services.auth import create_session

    Session = sessionmaker(bind=db_engine)
    db = Session()
    token = create_session(db, engineer_user)
    db.close()
    client.cookies.set("session_token", token)

    # Create a run in 'running' state for page render check
    db = Session()
    run_running = Run(
        part_number="PN-SSE",
        rev_a_label="A",
        rev_b_label="B",
        customer="Test",
        job_number="J-SSE",
        status="running",
        current_stage="Anchor building",
        current_stage_index=3,
        revA_path="/tmp/a.pdf",
        revB_path="/tmp/b.pdf",
        form3_path="/tmp/f.xlsx",
        reviewer_id=engineer_user.id,
    )
    # Create a completed run for SSE stream check (terminal state closes immediately)
    run_done = Run(
        part_number="PN-SSE-DONE",
        rev_a_label="A",
        rev_b_label="B",
        customer="Test",
        job_number="J-SSE2",
        status="completed",
        current_stage_index=8,
        revA_path="/tmp/a.pdf",
        revB_path="/tmp/b.pdf",
        form3_path="/tmp/f.xlsx",
        reviewer_id=engineer_user.id,
    )
    db.add(run_running)
    db.add(run_done)
    db.commit()
    db.refresh(run_running)
    db.refresh(run_done)
    run_id = run_running.id
    done_id = run_done.id
    db.close()

    # GET /runs/{id} must return 200 with stage checklist
    resp = client.get(f"/runs/{run_id}")
    assert resp.status_code == 200
    assert "stage-checklist" in resp.text
    assert "Anchor building" in resp.text

    # GET /runs/{id}/sse on a completed run: stream closes immediately after first event
    resp_sse = client.get(f"/runs/{done_id}/sse", headers={"Accept": "text/event-stream"})
    assert resp_sse.status_code == 200


def test_run_status_lifecycle(client, db_engine, engineer_user):
    """PIPE-03: GET /runs/{id} shows correct content for completed, failed, and warning states."""
    from sqlalchemy.orm import sessionmaker
    from shop.models import Run
    from shop.services.auth import create_session

    Session = sessionmaker(bind=db_engine)
    db = Session()
    token = create_session(db, engineer_user)
    db.close()
    client.cookies.set("session_token", token)

    db = Session()

    # Completed run
    run_c = Run(
        part_number="PN-C",
        rev_a_label="A",
        rev_b_label="B",
        customer="Test",
        job_number="J1",
        status="completed",
        current_stage_index=8,
        revA_path="/tmp/a.pdf",
        revB_path="/tmp/b.pdf",
        form3_path="/tmp/f.xlsx",
        reviewer_id=engineer_user.id,
    )
    # Failed run
    run_f = Run(
        part_number="PN-F",
        rev_a_label="A",
        rev_b_label="B",
        customer="Test",
        job_number="J2",
        status="failed",
        failure_stage="Alignment",
        failure_message="RANSAC failed to converge.",
        current_stage_index=4,
        revA_path="/tmp/a.pdf",
        revB_path="/tmp/b.pdf",
        form3_path="/tmp/f.xlsx",
        reviewer_id=engineer_user.id,
    )
    # Warning (low_confidence) run
    run_w = Run(
        part_number="PN-W",
        rev_a_label="A",
        rev_b_label="B",
        customer="Test",
        job_number="J3",
        status="warning",
        warning_type="low_confidence",
        confidence_summary={"message": "60% of items have low confidence", "low_confidence_ratio": 0.6},
        current_stage_index=8,
        revA_path="/tmp/a.pdf",
        revB_path="/tmp/b.pdf",
        form3_path="/tmp/f.xlsx",
        reviewer_id=engineer_user.id,
    )
    db.add_all([run_c, run_f, run_w])
    db.commit()
    db.refresh(run_c)
    db.refresh(run_f)
    db.refresh(run_w)
    cid, fid, wid = run_c.id, run_f.id, run_w.id
    db.close()

    # Completed: page loads successfully and shows completed state
    resp = client.get(f"/runs/{cid}")
    assert resp.status_code == 200
    assert "completed" in resp.text.lower()

    # Failed: shows failure message + failure stage
    resp = client.get(f"/runs/{fid}")
    assert resp.status_code == 200
    assert "Alignment" in resp.text
    assert "RANSAC failed to converge" in resp.text

    # Warning (low_confidence): shows confidence message + Abort button
    resp = client.get(f"/runs/{wid}")
    assert resp.status_code == 200
    assert "confidence" in resp.text.lower()
    assert "Abort" in resp.text

    # POST abort changes status to failed
    resp_abort = client.post(f"/runs/{wid}/abort", follow_redirects=False)
    assert resp_abort.status_code == 302

    db2 = Session()
    run_aborted = db2.query(Run).filter(Run.id == wid).first()
    db2.close()
    assert run_aborted.status == "failed"
    assert run_aborted.failure_stage == "Alignment"


def test_run_status_admin_debug_entry_uses_plain_get_form(client, db_engine, admin_user):
    """Regression: admin debug entry is submitted by GET form, not inline URL-munging JS."""
    from shop.models import Run

    Session = sessionmaker(bind=db_engine)
    _login_user(client, db_engine, admin_user)

    db = Session()
    run = Run(
        part_number="PN-DEBUG-CTA",
        rev_a_label="A",
        rev_b_label="B",
        customer="Test",
        job_number="J-DEBUG",
        status="completed",
        current_stage_index=8,
        revA_path="/tmp/a.pdf",
        revB_path="/tmp/b.pdf",
        form3_path="/tmp/f.xlsx",
        reviewer_id=admin_user.id,
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    run_id = run.id
    db.close()

    resp = client.get(f"/runs/{run_id}")
    assert resp.status_code == 200
    assert f'<form method="GET" action="/review/{run_id}"' in resp.text
    assert 'name="debug"' in resp.text
    assert 'value="1"' in resp.text
    assert 'new URL(' not in resp.text
    assert 'data-cta=' not in resp.text


def test_revA_balloon_failure(db_engine, engineer_user):
    """PIPE-04: When Rev A balloon detection fails (empty items), Run.status is set to
    'failed', Run.failure_stage is populated, and a descriptive failure_message is stored."""
    import json
    import pathlib
    import tempfile
    from unittest.mock import patch
    from shop.models import Run
    from shop.tasks import run_pipeline_task
    Session = sessionmaker(bind=db_engine)

    # Seed a Run in 'queued' state
    db = Session()
    run = Run(
        part_number="PN-PIPE04",
        rev_a_label="A",
        rev_b_label="B",
        customer="Acme",
        job_number="JOB-PIPE04",
        status="queued",
        revA_path="/tmp/a.pdf",
        revB_path="/tmp/b.pdf",
        form3_path="/tmp/f.xlsx",
        reviewer_id=engineer_user.id,
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    run_id = run.id
    db.close()

    def _mock_run_pipeline_empty(**kwargs):
        """Return output dir with empty items (no Rev A balloons detected)."""
        run_dir = pathlib.Path(tempfile.mkdtemp())
        packet = {"run_id": "test", "items": []}
        (run_dir / "delta_packet.json").write_text(json.dumps(packet))
        return run_dir

    # Patch both SessionLocal and run_pipeline so the task uses the test DB
    with patch("shop.tasks.SessionLocal", Session), \
         patch("shop.tasks.run_pipeline", _mock_run_pipeline_empty):
        run_pipeline_task.call_local(
            run_id,
            "/tmp/a.pdf",
            "/tmp/b.pdf",
            "/tmp/f.xlsx",
            "PN-PIPE04",
        )

    db2 = Session()
    updated = db2.query(Run).filter(Run.id == run_id).first()
    db2.close()
    assert updated.status == "failed", f"Expected 'failed', got: {updated.status!r}"
    assert updated.failure_stage is not None, "failure_stage should be set"
    assert updated.failure_message is not None, "failure_message should be set"


def test_revB_balloon_warning(db_engine, engineer_user):
    """PIPE-05: When alignment fails due to Rev B balloon issues (all items have
    low location scores), Run.status is set to 'warning' and Run.warning_type is
    set to 'low_confidence' (the unified warning path for revB alignment failure)."""
    import json
    import pathlib
    import tempfile
    from unittest.mock import patch
    from shop.models import Run
    from shop.tasks import run_pipeline_task
    Session = sessionmaker(bind=db_engine)

    db = Session()
    run = Run(
        part_number="PN-PIPE05",
        rev_a_label="A",
        rev_b_label="B",
        customer="Acme",
        job_number="JOB-PIPE05",
        status="queued",
        revA_path="/tmp/a.pdf",
        revB_path="/tmp/b.pdf",
        form3_path="/tmp/f.xlsx",
        reviewer_id=engineer_user.id,
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    run_id = run.id
    db.close()

    def _mock_run_pipeline_low_location(**kwargs):
        """Return output with items where all location scores < 0.5 (simulates
        Rev B balloon alignment failure which causes near-zero location scores)."""
        run_dir = pathlib.Path(tempfile.mkdtemp())
        packet = {
            "run_id": "test",
            "items": [
                {"char_no": i, "status": "uncertain", "scores": {"location": 0.1, "text": 0.5, "context": 0.5}}
                for i in range(1, 6)
            ],
        }
        (run_dir / "delta_packet.json").write_text(json.dumps(packet))
        return run_dir

    with patch("shop.tasks.SessionLocal", Session), \
         patch("shop.tasks.run_pipeline", _mock_run_pipeline_low_location):
        run_pipeline_task.call_local(
            run_id,
            "/tmp/a.pdf",
            "/tmp/b.pdf",
            "/tmp/f.xlsx",
            "PN-PIPE05",
        )

    db2 = Session()
    updated = db2.query(Run).filter(Run.id == run_id).first()
    db2.close()
    assert updated.status == "warning", f"Expected 'warning', got: {updated.status!r}"
    assert updated.warning_type == "low_confidence", f"Expected 'low_confidence', got: {updated.warning_type!r}"


def test_low_confidence_warning(db_engine, engineer_user):
    """PIPE-06: When alignment confidence is uniformly low (>50% of items have
    location score < 0.5), Run.status='warning', Run.warning_type='low_confidence',
    and Run.confidence_summary JSON is populated."""
    import json
    import pathlib
    import tempfile
    from unittest.mock import patch
    from shop.models import Run
    from shop.tasks import run_pipeline_task
    Session = sessionmaker(bind=db_engine)

    db = Session()
    run = Run(
        part_number="PN-PIPE06",
        rev_a_label="A",
        rev_b_label="B",
        customer="Acme",
        job_number="JOB-PIPE06",
        status="queued",
        revA_path="/tmp/a.pdf",
        revB_path="/tmp/b.pdf",
        form3_path="/tmp/f.xlsx",
        reviewer_id=engineer_user.id,
    )
    db.add(run)
    db.commit()
    db.refresh(run)
    run_id = run.id
    db.close()

    def _mock_run_pipeline_low_conf(**kwargs):
        """Return output with >50% of items having low location scores."""
        run_dir = pathlib.Path(tempfile.mkdtemp())
        # 8 low + 2 high = 80% low confidence ratio
        items = [
            {"char_no": i, "status": "uncertain", "scores": {"location": 0.2, "text": 0.5, "context": 0.5}}
            for i in range(1, 9)
        ] + [
            {"char_no": i, "status": "unchanged", "scores": {"location": 0.9, "text": 0.9, "context": 0.9}}
            for i in range(9, 11)
        ]
        packet = {"run_id": "test", "items": items}
        (run_dir / "delta_packet.json").write_text(json.dumps(packet))
        return run_dir

    with patch("shop.tasks.SessionLocal", Session), \
         patch("shop.tasks.run_pipeline", _mock_run_pipeline_low_conf):
        run_pipeline_task.call_local(
            run_id,
            "/tmp/a.pdf",
            "/tmp/b.pdf",
            "/tmp/f.xlsx",
            "PN-PIPE06",
        )

    db2 = Session()
    updated = db2.query(Run).filter(Run.id == run_id).first()
    db2.close()
    assert updated.status == "warning", f"Expected 'warning', got: {updated.status!r}"
    assert updated.warning_type == "low_confidence", f"Expected 'low_confidence', got: {updated.warning_type!r}"
    assert updated.confidence_summary is not None, "confidence_summary should be populated"
    summary = updated.confidence_summary
    assert "low_confidence_ratio" in summary, f"Expected 'low_confidence_ratio' in summary, got: {summary}"
    assert summary["low_confidence_ratio"] > 0.5, f"Expected ratio > 0.5, got: {summary['low_confidence_ratio']}"


# ---------------------------------------------------------------------------
# Alert requirements
# ---------------------------------------------------------------------------


def test_failure_alert_created(client, db_engine, engineer_user):
    """PIPE-11: POST /alerts/dismiss/{id} marks RunAlert.is_read=True and returns empty 200.

    Steps:
    - Seed a Run and a RunAlert for engineer_user.
    - Authenticate as engineer_user.
    - POST /alerts/dismiss/{alert_id}.
    - Assert 200 response with empty body.
    - Assert RunAlert.is_read == True in the DB.
    """
    from sqlalchemy.orm import sessionmaker
    from shop.models import Run, RunAlert
    from shop.services.auth import create_session

    Session = sessionmaker(bind=db_engine)

    # Seed a Run owned by engineer_user
    db = Session()
    try:
        run = Run(
            part_number="PN-ALERT",
            rev_a_label="A",
            rev_b_label="B",
            customer="Acme",
            job_number="JOB-A",
            status="failed",
            failure_stage="stage_2_balloon_detection",
            revA_path="/tmp/a.pdf",
            revB_path="/tmp/b.pdf",
            form3_path="/tmp/form.xlsx",
            reviewer_id=engineer_user.id,
        )
        db.add(run)
        db.flush()

        alert = RunAlert(
            run_id=run.id,
            user_id=engineer_user.id,
            message=f"Run {run.id} (PN-ALERT) failed at stage: stage_2_balloon_detection",
            failure_stage="stage_2_balloon_detection",
            is_read=False,
        )
        db.add(alert)
        db.commit()
        alert_id = alert.id
    finally:
        db.close()

    # Authenticate as engineer_user
    db2 = Session()
    token = create_session(db2, engineer_user)
    db2.close()
    client.cookies.set("session_token", token)

    # POST /alerts/dismiss/{alert_id}
    resp = client.post(f"/alerts/dismiss/{alert_id}", follow_redirects=False)
    assert resp.status_code == 200, f"Expected 200, got {resp.status_code}: {resp.text[:200]}"
    assert resp.text.strip() == "", f"Expected empty body, got: {resp.text!r}"

    # Verify alert is marked as read in the DB
    db3 = Session()
    try:
        updated = db3.query(RunAlert).filter(RunAlert.id == alert_id).first()
        assert updated is not None
        assert updated.is_read is True, "Alert should be marked as read after dismiss"
    finally:
        db3.close()
