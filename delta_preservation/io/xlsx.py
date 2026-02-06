"""
AS9102 Form 3 parsing utilities for extracting inspection characteristics from Excel files.

This module handles the parsing of AS9102 Form 3 inspection documents, which contain
detailed characteristic information including requirements, reference locations, and
characteristic designators. The module provides robust column detection and fallback
mechanisms to handle various Form 3 formats.
"""

from pathlib import Path
from typing import List
import json
import html
from openpyxl import load_workbook


class Characteristic:
    """
    Represents a single row from AS9102 Form 3.
    
    Attributes:
        char_no: Integer characteristic number (e.g. 1, 2, 3...)
        reference_location: Reference location string from Form 3
        characteristic_designator: Characteristic designator string
        requirement: Requirement text string
    """
    def __init__(self, char_no: int, reference_location: str, 
                 characteristic_designator: str, requirement: str):
        self.char_no = char_no
        self.reference_location = reference_location
        self.characteristic_designator = characteristic_designator
        self.requirement = requirement
    
    def to_dict(self) -> dict:
        """
        Convert characteristic to dictionary for JSON serialization.
        
        Returns:
            Dictionary representation of the characteristic with all fields
            as key-value pairs suitable for JSON serialization or debugging output.
        """
        return {
            "char_no": self.char_no,
            "reference_location": self.reference_location,
            "characteristic_designator": self.characteristic_designator,
            "requirement": self.requirement
        }


def load_form3(xlsx_path: Path, intermediate_dir: Path) -> List[Characteristic]:
    """
    Parse AS9102 Form 3 from an Excel file and extract characteristic rows.
    
    This function:
    1. Opens the workbook and selects the "Form3" or "F3" worksheet
    2. Scans all rows to identify the header row containing column positions for:
       - Char no
       - Reference location
       - Characteristic Designator
       - Requirement
    3. Uses keyword matching to identify headers (requires at least 3 matching columns)
    4. Raises an error if no valid header row is found or if required columns are missing
    5. Iterates from the row immediately after headers until Char no cell is empty
    6. Parses each row into a Characteristic object with minimal cleanup:
       - Char no: parsed as integer
       - Other fields: stripped strings (empty string if blank/None)
    7. Writes raw parsed data to intermediate/form3_chars.json for debugging
    8. Returns the list of Characteristic objects
    
    Args:
        xlsx_path: Path to the Form 3 Excel file
        intermediate_dir: Directory for debug output (form3_chars.json)
    
    Returns:
        List of Characteristic objects, one per valid row starting after headers
    
    Raises:
        FileNotFoundError: If xlsx_path does not exist
        ValueError: If neither "Form3" nor "F3" worksheet is found
        ValueError: If no valid header row is found or required columns are missing
    """
    # Load workbook with data_only=True to get calculated values instead of formulas
    wb = load_workbook(xlsx_path, data_only=True)
    
    # Search for the Form 3 worksheet using flexible naming
    # AS9102 forms may have variations like "Form3", "F3", "Form 3", etc.
    sheet_name = None
    for name in wb.sheetnames:
        if "Form3" in name or "F3" in name:
            sheet_name = name
            break
    
    if sheet_name is None:
        raise ValueError(f"Neither 'Form3' nor 'F3' worksheet found in {xlsx_path}")
    
    ws = wb[sheet_name]
    
    # Attempt intelligent column detection by scanning all rows for headers
    # This allows the parser to adapt to Form 3 variations and custom layouts
    column_map = {}
    header_keywords = {
        "char_no": ["char", "no", "number"],
        "reference_location": ["reference", "location", "ref"],
        "characteristic_designator": ["characteristic", "designator"],
        "requirement": ["requirement", "req"]
    }
    
    header_row_idx = None
    
    # Scan all rows to find the header row
    for row_idx in range(1, ws.max_row + 1):
        temp_column_map = {}
        
        # Scan current row for potential headers
        for col_idx in range(1, ws.max_column + 1):
            cell_value = ws.cell(row=row_idx, column=col_idx).value
            if cell_value is None:
                continue
            
            cell_text = str(cell_value).lower().strip()
            
            # Check each field we're looking for
            for field_name, keywords in header_keywords.items():
                if field_name not in temp_column_map:
                    # Apply field-specific matching logic
                    if field_name == "char_no":
                        # Require "char" AND ("no" OR "number")
                        if "char" in cell_text and ("no" in cell_text or "number" in cell_text):
                            temp_column_map[field_name] = col_idx
                    elif field_name == "requirement":
                        # Require "requirement" OR "req"
                        if "requirement" in cell_text or "req" in cell_text:
                            temp_column_map[field_name] = col_idx
                    elif field_name == "reference_location":
                        # Require "ref" OR "reference"
                        if "ref" in cell_text or "reference" in cell_text:
                            temp_column_map[field_name] = col_idx
                    elif field_name == "characteristic_designator":
                        # Require "characteristic" OR "designator"
                        if "characteristic" in cell_text or "designator" in cell_text:
                            temp_column_map[field_name] = col_idx
        
        # If we found at least 3 matching columns, consider this the header row
        if len(temp_column_map) >= 3:
            column_map = temp_column_map
            header_row_idx = row_idx
            break
    
    # If no header row found, raise an error
    if header_row_idx is None:
        raise ValueError(f"No valid header row found in {xlsx_path}. Could not locate AS9102 Form 3 headers.")
    
    # Ensure all required columns are found
    required_fields = ["char_no", "reference_location", "characteristic_designator", "requirement"]
    missing_fields = [field for field in required_fields if field not in column_map]
    if missing_fields:
        raise ValueError(f"Missing required columns in header row {header_row_idx}: {missing_fields}")
    
    # Parse characteristics starting immediately after the header row
    characteristics = []
    row_idx = header_row_idx + 1
    
    while True:
        # Read Char no cell
        char_no_cell = ws.cell(row=row_idx, column=column_map["char_no"]).value
        
        # Stop if Char no is empty/None
        if char_no_cell is None or str(char_no_cell).strip() == "":
            break
        
        # Parse Char no as integer
        try:
            char_no = int(char_no_cell)
        except (ValueError, TypeError):
            # Skip rows where Char no cannot be parsed as integer
            row_idx += 1
            continue
        
        # Read other fields and coerce to stripped strings
        ref_loc_cell = ws.cell(row=row_idx, column=column_map["reference_location"]).value
        reference_location = str(ref_loc_cell).strip() if ref_loc_cell is not None else ""
        
        char_des_cell = ws.cell(row=row_idx, column=column_map["characteristic_designator"]).value
        characteristic_designator = str(char_des_cell).strip() if char_des_cell is not None else ""
        
        req_cell = ws.cell(row=row_idx, column=column_map["requirement"]).value
        requirement = str(req_cell).strip() if req_cell is not None else ""
        
        # Create Characteristic object
        char = Characteristic(
            char_no=char_no,
            reference_location=reference_location,
            characteristic_designator=characteristic_designator,
            requirement=requirement
        )
        characteristics.append(char)
        
        row_idx += 1
    
    # Serialize to JSON for debugging
    debug_output = [char.to_dict() for char in characteristics]
    debug_path = intermediate_dir / "form3_chars.json"
    with open(debug_path, "w", encoding="utf-8") as f:
        json.dump(debug_output, f, indent=2, ensure_ascii=False)
    
    return characteristics
